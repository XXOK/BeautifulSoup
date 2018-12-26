import re
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

def get_html(url):
   _html = ""
   resp = requests.get(url)
   if resp.status_code == 200:
      _html = resp.text
   return _html

URL = "https://www.myrealtrip.com/experiences"
html = get_html(URL)
soup = BeautifulSoup(html, 'html.parser')

items_name = soup.find_all('div',{'class':'name'}, limit=12)

items_price = soup.find_all(class_='price', limit=12)

items_url = soup.find_all("a", href=re.compile("^(/offers/)"), limit=12)

wb = Workbook()
ws = wb.active
ws.title = '인기 가이드투어'

count = 1
# 엑셀 셀 인덱스
url = []
# 상품 url 빈 리스트
url_count = 0
# 상품 url 리스트 인덱스

for i,j,k in zip(items_name, items_price, items_url):

      price = j.get_text()
      price_split1 = price.split('원')
      price_split2 = price_split1[0].splitlines() # splitlines -> 자동 줄바꿈 문자 제거

      ws['A'+str(count)] = str(i.get_text())
      ws['B'+str(count)] = str(price_split2[1]+'원')

      for link in items_url:
          if 'href' in link.attrs:
              url.append(link.attrs['href'])
      ws['C' + str(count)] = str('https://www.myrealtrip.com' + url[url_count])

      print(ws['A' + str(count)].value, ws['B' + str(count)].value, ws['C' + str(count)].value)

      url_count += 1
      count += 1

wb.save('/Users/yeonshin/BeautifulSoup/인기 가이드투어.xlsx')